# # https://github.com/meer-khan/Art-Gallery-App-Flutter/commits/main


# import requests, re
# url = "https://github.com/ZhangGongjie/thesis/commits/master"
# # url = "https://github.com/ZhangGongjie/ZhangGongjie.github.io/commits/master"
# text = requests.get(url=url).text

# print(text)

# pattern = re.compile(r"\w+\s+\d{1,2},\s+\d{3,4}", re.I|re.M)
# print(pattern)
# match = pattern.findall(text)
# print(match)
# # Commits on Mar 31, 2022



from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active 
fake_data = Faker()

for i in range(1,11):
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()
    ws.cell(row=i, column=3).value = fake_data.address()


wb.save("testData.csv")